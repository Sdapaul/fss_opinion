"""SQLite 기반 이메일 발송 이력 저장소 및 엑셀 export."""

import io
import sqlite3
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

DB_PATH = Path("fss_items.db")

_CREATE_SQL = """
CREATE TABLE IF NOT EXISTS items (
    id        TEXT PRIMARY KEY,
    title     TEXT NOT NULL,
    date      TEXT NOT NULL,
    url       TEXT NOT NULL,
    category  TEXT NOT NULL,
    field     TEXT,
    serial_no TEXT,
    summary   TEXT,
    sent_at   TEXT NOT NULL
)
"""

# 기존 테이블에 컬럼이 없을 경우 추가 (무시 가능한 실패)
_MIGRATE_SQLS = [
    "ALTER TABLE items ADD COLUMN field TEXT",
    "ALTER TABLE items ADD COLUMN serial_no TEXT",
]


def _conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.execute(_CREATE_SQL)
    conn.commit()
    # 마이그레이션: 이미 있는 컬럼은 오류 무시
    for sql in _MIGRATE_SQLS:
        try:
            conn.execute(sql)
            conn.commit()
        except sqlite3.OperationalError:
            pass
    return conn


def init_db() -> None:
    """DB 파일과 테이블을 미리 생성."""
    _conn().close()


def save_items(items: list[dict]) -> None:
    """발송된 항목을 DB에 저장 (중복 무시)."""
    today = date.today().isoformat()
    with _conn() as conn:
        conn.executemany(
            """INSERT OR IGNORE INTO items
               (id, title, date, url, category, field, serial_no, summary, sent_at)
               VALUES (:id, :title, :date, :url, :category, :field, :serial_no, :summary, :sent_at)""",
            [
                {
                    **item,
                    "field": item.get("field", ""),
                    "serial_no": item.get("serial_no", ""),
                    "summary": item.get("summary", ""),
                    "sent_at": today,
                }
                for item in items
            ],
        )


def query_items(date_from: str, date_to: str) -> list[dict]:
    """기간별 발송 항목 조회 (date: YYYY-MM-DD)."""
    with _conn() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM items WHERE date BETWEEN ? AND ? ORDER BY date DESC, category",
            (date_from, date_to),
        ).fetchall()
    return [dict(row) for row in rows]


def export_excel_bytes(date_from: str, date_to: str) -> bytes:
    """기간별 항목을 엑셀 바이트로 반환."""
    rows = query_items(date_from, date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FSS 알림 이력"

    headers = ["구분", "분야", "일련번호", "제목", "등록일", "발송일", "AI요약", "원문 URL"]
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, item in enumerate(rows, 2):
        ws.cell(r, 1, item["category"])
        ws.cell(r, 2, item.get("field", ""))
        ws.cell(r, 3, item.get("serial_no", ""))
        ws.cell(r, 4, item["title"]).alignment = Alignment(wrap_text=True)
        ws.cell(r, 5, item["date"])
        ws.cell(r, 6, item["sent_at"])
        ws.cell(r, 7, item.get("summary", "")).alignment = Alignment(wrap_text=True)
        url_cell = ws.cell(r, 8, item["url"])
        url_cell.hyperlink = item["url"]
        url_cell.font = Font(color="1D4ED8", underline="single")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 55
    ws.column_dimensions["H"].width = 80
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
